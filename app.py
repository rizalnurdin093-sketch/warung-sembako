"""Warung Sembako — Fase 2: Login role + Anti-kecurangan. Flask + SQLite."""
from flask import Flask, render_template, request, redirect, url_for, jsonify, flash
from functools import wraps
import database as db

app = Flask(__name__)
app.secret_key = 'warung-sembako-rahasia'
# Mounted di subpath /warung via WSGI DispatcherMiddleware (Nginx proxy)
app.config['APPLICATION_ROOT'] = '/warung'

# ===== Auth helpers =====
def get_user(username):
    return db.user_find(username)

def login_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if 'user_id' not in request.cookies:
            return redirect(url_for('login'))
        return f(*args, **kwargs)
    return decorated

def role_required(*roles):
    def decorator(f):
        @wraps(f)
        def decorated(*args, **kwargs):
            if 'role' not in request.cookies:
                return redirect(url_for('login'))
            if request.cookies.get('role') not in roles:
                flash('Akses ditolak: hanya untuk ' + ', '.join(roles), 'error')
                return redirect(url_for('kasir'))
            return f(*args, **kwargs)
        return decorated
    return decorator

def current_user():
    """Return current user info from cookies."""
    if 'user_id' in request.cookies:
        return {
            'id': int(request.cookies.get('user_id')),
            'username': request.cookies.get('username'),
            'role': request.cookies.get('role'),
            'nama': request.cookies.get('nama'),
        }
    return None

@app.context_processor
def inject_user():
    return dict(current_user=current_user())

@app.context_processor
def inject_stok_rendah():
    """Available di semua template: stok_rendah_count, stok_rendah_produk."""
    if 'user_id' in request.cookies and request.cookies.get('role') == 'owner':
        r = db.stok_rendah()
        return dict(stok_rendah_count=len(r), stok_rendah_produk=r)
    return dict(stok_rendah_count=0, stok_rendah_produk=[])

# ===== LOGIN / LOGOUT =====
@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        username = request.form['username'].strip()
        password = request.form['password'].strip()
        user = get_user(username)
        if user and user['password'] == password:  # plain password untuk simpilitas Fase 2
            resp = redirect(url_for('kasir'))
            resp.set_cookie('user_id', str(user['id']), max_age=86400*30, httponly=True)
            resp.set_cookie('username', user['username'], max_age=86400*30, httponly=True)
            resp.set_cookie('role', user['role'], max_age=86400*30, httponly=True)
            resp.set_cookie('nama', user['nama'] or '', max_age=86400*30, httponly=True)
            return resp
        flash('Username/password salah', 'error')
    return render_template('login.html')

@app.route('/logout')
def logout():
    resp = redirect(url_for('login'))
    for k in ['user_id', 'username', 'role', 'nama']:
        resp.set_cookie(k, '', expires=0)
    return resp

# ===== KASIR (transaksi jual) — BISA OWNER & KASIR =====
@app.route('/')
@login_required
def index():
    return redirect(url_for('kasir'))

@app.route('/kasir')
@login_required
def kasir():
    produk = db.produk_list()
    return render_template('kasir.html', produk=produk)

@app.route('/api/produk/<int:pid>')
@login_required
def api_produk(pid):
    produk = db.produk_list()
    for p in produk:
        if p['id'] == pid:
            return jsonify(dict(p))
    return jsonify({'error': 'not found'}), 404

@app.route('/jual', methods=['POST'])
@login_required
def jual():
    pid = int(request.form['produk_id'])
    qty = int(request.form['qty'])
    produk = db.produk_list()
    p = next((x for x in produk if x['id'] == pid), None)
    if not p:
        return 'Produk tidak ditemukan', 400
    if qty < 1:
        return 'Qty minimal 1', 400
    if qty > p['stok']:
        return f'Stok tidak cukup (sisa {p["stok"]})', 400
    kasir = current_user()['nama'] or current_user()['username']
    db.transaksi_add(p['nama'], p['harga_jual'], qty, kasir)
    db.stok_kurang(pid, qty)
    flash('Penjualan tercatat', 'success')
    return redirect(url_for('kasir'))

# ===== REKAP — OWNER ONLY =====
@app.route('/rekap')
@login_required
@role_required('owner')
def rekap():
    rows = db.rekap_harian()
    total_all = sum(r['total'] for r in rows)
    return render_template('rekap.html', rows=rows, total_all=total_all)

# ===== PRODUK — OWNER ONLY =====
@app.route('/produk')
@login_required
@role_required('owner')
def produk():
    produk_rows = db.produk_list()
    return render_template('produk.html', produk=produk_rows)

@app.route('/produk/tambah', methods=['POST'])
@login_required
@role_required('owner')
def produk_tambah():
    db.produk_add(
        request.form['nama'],
        request.form['satuan'],
        int(request.form['harga_beli'] or 0),
        int(request.form['harga_jual'] or 0),
        int(request.form['stok'] or 0),
    )
    flash('Produk ditambahkan', 'success')
    return redirect(url_for('produk'))

@app.route('/produk/hapus/<int:pid>')
@login_required
@role_required('owner')
def produk_hapus(pid):
    db.produk_delete(pid)
    flash('Produk dihapus', 'success')
    return redirect(url_for('produk'))

# ===== SETORAN KAS (anti-kecurangan) — OWNER =====
@app.route('/setoran')
@login_required
@role_required('owner')
def setoran():
    rows = db.setoran_list()
    return render_template('setoran.html', rows=rows)

@app.route('/setoran/tambah', methods=['POST'])
@login_required
@role_required('owner')
def setoran_tambah():
    db.setoran_add(request.form['tanggal'], int(request.form['jumlah']), request.form['catatan'])
    flash('Setoran dicatat', 'success')
    return redirect(url_for('setoran'))

# ===== OPNAME STOK (anti-kecurangan) — OWNER =====
@app.route('/opname')
@login_required
@role_required('owner')
def opname():
    rows = db.opname_list()
    produk = db.produk_list()
    return render_template('opname.html', rows=rows, produk=produk)

@app.route('/opname/tambah', methods=['POST'])
@login_required
@role_required('owner')
def opname_tambah():
    pid = int(request.form['produk_id'])
    stok_fisik = int(request.form['stok_fisik'])
    # ambil stok sistem
    p = next((x for x in db.produk_list() if x['id'] == pid), None)
    stok_sistem = p['stok'] if p else 0
    db.opname_add(request.form['tanggal'], pid, stok_fisik, stok_sistem, request.form['catatan'])
    flash('Opname dicatat', 'success')
    return redirect(url_for('opname'))

# ===== DASHBOARD OWNER (selisih kas + stok) =====
@app.route('/dashboard')
@login_required
@role_required('owner')
def dashboard():
    # ambil 7 hari terakhir
    conn = db.get_db()
    hari_rows = conn.execute("""
        SELECT date(created_at) as hari FROM transaksi
        GROUP BY date(created_at)
        ORDER BY hari DESC LIMIT 7
    """).fetchall()
    conn.close()
    
    dashboard_data = []
    for h in hari_rows:
        tanggal = h['hari']
        kas_teoretis = db.omzet_by_date(tanggal)
        kas_setoran = db.setoran_today(tanggal)
        selisih_kas = kas_setoran - kas_teoretis
        
        # hitung total selisih stok per hari (dari opname)
        conn = db.get_db()
        stok_row = conn.execute("""
            SELECT SUM(selisih) as total_selisih FROM opname WHERE tanggal=?
        """, (tanggal,)).fetchone()
        conn.close()
        selisih_stok = stok_row['total_selisih'] or 0
        
        dashboard_data.append({
            'tanggal': tanggal,
            'kas_teoretis': kas_teoretis,
            'kas_setoran': kas_setoran,
            'selisih_kas': selisih_kas,
            'selisih_stok': selisih_stok
        })
    
    # total_all = total omzet (untuk header) — dari kas teoretis semua hari
    total_all = sum(d['kas_teoretis'] for d in dashboard_data)
    total_tx = sum(r['total'] for r in db.grafik_omzet_harian(365))
    jml_produk = len(db.produk_list())
    jml_kasir = len([u for u in db.user_list() if u['role'] == 'kasir'])
    # data grafik omzet 30 hari + per produk
    omzet = db.grafik_omzet_harian(30)
    prod = db.grafik_penjualan_produk()
    grafik_labels = [r['tgl'] for r in omzet]
    grafik_totals = [r['total'] for r in omzet]
    g_prod_labels = [r['nama'] for r in prod]
    g_prod_vals = [r['total'] for r in prod]
    return render_template('dashboard.html',
        data=dashboard_data, total_all=total_all,
        total_tx=total_tx, jml_produk=jml_produk, jml_kasir=jml_kasir,
        grafik_labels=grafik_labels, grafik_totals=grafik_totals,
        g_prod_labels=g_prod_labels, g_prod_vals=g_prod_vals)

# ===== EXPORT CSV/EXCEL =====
import csv
import io
from datetime import datetime

@app.route('/export/csv/<jenis>')
@login_required
@role_required('owner')
def export_csv(jenis):
    """Export data ke CSV. jenis: transaksi, produk, setoran, opname"""
    output = io.StringIO()
    writer = csv.writer(output)
    
    if jenis == 'transaksi':
        writer.writerow(['ID', 'Tanggal', 'Produk', 'Harga', 'Qty', 'Total', 'Kasir'])
        rows = db.get_db().execute("""
            SELECT id, created_at, nama_produk, harga_satuan, qty, total, kasir
            FROM transaksi ORDER BY created_at DESC
        """).fetchall()
        for r in rows:
            writer.writerow([r['id'], r['created_at'], r['nama_produk'], 
                           r['harga_satuan'], r['qty'], r['total'], r['kasir'] or ''])
        filename = f'export_transaksi_{datetime.now().strftime("%Y%m%d")}.csv'
    
    elif jenis == 'produk':
        writer.writerow(['ID', 'Nama', 'Satuan', 'Harga Beli', 'Harga Jual', 'Stok'])
        rows = db.produk_list()
        for r in rows:
            writer.writerow([r['id'], r['nama'], r['satuan'], r['harga_beli'], r['harga_jual'], r['stok']])
        filename = f'export_produk_{datetime.now().strftime("%Y%m%d")}.csv'
    
    elif jenis == 'setoran':
        writer.writerow(['ID', 'Tanggal', 'Jumlah', 'Catatan', 'Waktu Input'])
        rows = db.setoran_list()
        for r in rows:
            writer.writerow([r['id'], r['tanggal'], r['jumlah'], r['catatan'] or '', r['created_at']])
        filename = f'export_setoran_{datetime.now().strftime("%Y%m%d")}.csv'
    
    elif jenis == 'opname':
        writer.writerow(['ID', 'Tanggal', 'Produk', 'Stok Sistem', 'Stok Fisik', 'Selisih', 'Catatan', 'Waktu Input'])
        rows = db.opname_list()
        for r in rows:
            writer.writerow([r['id'], r['tanggal'], r['nama_produk'], r['stok_sistem'], 
                           r['stok_fisik'], r['selisih'], r['catatan'] or '', r['created_at']])
        filename = f'export_opname_{datetime.now().strftime("%Y%m%d")}.csv'
    
    else:
        return 'Jenis export tidak dikenal', 400
    
    response = output.getvalue()
    return response, 200, {
        'Content-Type': 'text/csv; charset=utf-8',
        'Content-Disposition': f'attachment; filename="{filename}"'
    }

# ===== EXPORT EXCEL (.xlsx) =====
from io import BytesIO

def _xlsx_response(wb, filename):
    """Convert workbook ke HTTP response."""
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue(), 200, {
        'Content-Type': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        'Content-Disposition': f'attachment; filename="{filename}"'
    }

@app.route('/export/xlsx/<jenis>')
@login_required
@role_required('owner')
def export_xlsx(jenis):
    """Export data ke Excel (.xlsx) — sama seperti CSV tapi .xlsx."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill

    wb = openpyxl.Workbook()
    ws = wb.active
    # styling header
    header_fill = PatternFill(start_color='2D6A4F', end_color='2D6A4F', fill_type='solid')
    header_font = Font(color='FFFFFF', bold=True)

    def write_rows(headers, rows):
        ws.append(headers)
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
        for r in rows:
            ws.append(r)

    if jenis == 'transaksi':
        write_rows(['ID', 'Tanggal', 'Produk', 'Harga', 'Qty', 'Total', 'Kasir'],
            [ [r['id'], r['created_at'], r['nama_produk'], r['harga_satuan'], r['qty'], r['total'], r['kasir'] or '']
              for r in db.get_db().execute("SELECT * FROM transaksi ORDER BY created_at DESC").fetchall() ])
        filename = f'warung_transaksi.xlsx'

    elif jenis == 'produk':
        write_rows(['ID', 'Nama', 'Satuan', 'Harga Beli', 'Harga Jual', 'Stok'],
            [ [r['id'], r['nama'], r['satuan'], r['harga_beli'], r['harga_jual'], r['stok']]
              for r in db.produk_list() ])
        filename = f'warung_produk.xlsx'

    elif jenis == 'setoran':
        write_rows(['ID', 'Tanggal', 'Jumlah', 'Catatan', 'Waktu Input'],
            [ [r['id'], r['tanggal'], r['jumlah'], r['catatan'] or '', r['created_at']]
              for r in db.setoran_list() ])
        filename = f'warung_setoran.xlsx'

    elif jenis == 'opname':
        write_rows(['ID', 'Tanggal', 'Produk', 'Stok Sistem', 'Stok Fisik', 'Selisih', 'Catatan', 'Waktu Input'],
            [ [r['id'], r['tanggal'], r['nama_produk'], r['stok_sistem'], r['stok_fisik'], r['selisih'], r['catatan'] or '', r['created_at']]
              for r in db.opname_list() ])
        filename = f'warung_opname.xlsx'

    else:
        return 'Jenis export tidak dikenal', 400

    return _xlsx_response(wb, filename)

# ===== AUDIT LOG (histori transaksi detail per kasir) =====
@app.route('/audit')
@login_required
@role_required('owner')
def audit():
    transaksi = db.transaksi_all()
    per_kasir = db.rekap_per_kasir()
    total_all = sum(r['total'] for r in transaksi)
    return render_template('audit.html', transaksi=transaksi, per_kasir=per_kasir, total_all=total_all)

# ===== BACKUP / RESTORE DATABASE =====
@app.route('/backup')
@login_required
@role_required('owner')
def backup_db():
    """Download backup database."""
    data = db.backup_blob()
    tgl = datetime.now().strftime("%Y%m%d_%H%M")
    return data, 200, {
        'Content-Type': 'application/octet-stream',
        'Content-Disposition': f'attachment; filename="warung_backup_{tgl}.db"'
    }

@app.route('/restore', methods=['GET', 'POST'])
@login_required
@role_required('owner')
def restore_db():
    if request.method == 'POST':
        f = request.files.get('file')
        if not f or f.filename == '':
            flash('Pilih file backup dulu', 'error')
            return redirect(url_for('restore_db'))
        try:
            data = f.read()
            db.restore_blob(data)
            flash('Database berhasil di-restore', 'success')
            return redirect(url_for('dashboard'))
        except Exception as e:
            flash(f'Gagal restore: {e}', 'error')
            return redirect(url_for('restore_db'))
    return render_template('restore.html')

# ===== GRAFIK PENJUALAN =====
@app.route('/grafik')
@login_required
@role_required('owner')
def grafik():
    omzet = db.grafik_omzet_harian(30)
    produk = db.grafik_penjualan_produk()
    labels = [r['tgl'] for r in omzet]
    totals = [r['total'] for r in omzet]
    p_labels = [r['nama'] for r in produk]
    p_vals = [r['total'] for r in produk]
    return render_template('grafik.html',
        labels=labels, totals=totals,
        p_labels=p_labels, p_vals=p_vals)

if __name__ == '__main__':
    app.run(host='0.0.0.0', port=5002, debug=True)